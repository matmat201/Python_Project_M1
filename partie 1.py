# -*- coding: utf-8 -*-
"""
Created on Thu Feb 22 15:55:02 2024

@author: mathi
"""

###############
#Import des bibliotheque
###############
import openpyxl as op
import matplotlib.pyplot as plt

##############
#Programme principal
##############

book = op.load_workbook('bdd_photons.xlsx')

## choix materiaux
list_materiaux = ["Aluminium", "Plomb", "Cobalt", "Cuivre"] #plutot que de repeter les instruction de plot et d'extraction on demande a l'utilisateur de choisir parmis la liste

while True: #on effectue un test pour verifier la valeur saisie
    materiaux = input("Choisissez le matériau (Aluminium, Plomb, Cobalt, Cuivre) : ").capitalize() #on utilise .capitalize pour s'assurer du bon formatage
    if materiaux in list_materiaux:
        break
    else:
        print("Matériau non valide. Veuillez choisir parmi Aluminium, Plomb, Cobalt ou Cuivre.")

sheet = book.get_sheet_by_name(materiaux) #on active la bonne page dans le document excel

## def var d'interet
Energie = []
Diff_c = []
Photoel = []
CP_nuc = []
CP_el = []

##extraction des donnée depuis excel
for i in range(4,sheet.max_row+1):                                  #on sait que les données commence a la 4ieme ligne
    Energie.append(float(sheet.cell(row = i, column = 1).value))
    Diff_c.append(float(sheet.cell(row = i, column = 2).value))
    Photoel.append(float(sheet.cell(row = i, column = 3).value))
    CP_nuc.append(float(sheet.cell(row = i, column = 4).value))
    CP_el.append(float(sheet.cell(row = i, column = 5).value))

##plot des données
plt.figure(dpi=200)    #augmente la resolution de l'image sinon le resultat n'est pas tres lisible dans le retour graphe
plt.title(f"Evolution coeff atténuation massique pour {materiaux}")    
plt.xlabel("Energie")
plt.ylabel("tau(cm2/g)")
plt.xscale('log')
plt.yscale('log')
plt.xlim(min(Energie),max(Energie)) #on set les limite de maniere dynamique en fonction de la liste de données choisie
plt.ylim(min(Photoel),max(Photoel))
plt.grid(which = "major", axis='y', linestyle = '--') #afin de s'assurer que le formatage de ligne et comme on le desire on affiche les centaine en Y et les dizaine en X
plt.grid(which = "both", axis='x', linestyle = '--')

plt.plot(Energie,Diff_c,label="Diffusion Compton")
plt.plot(Energie,Photoel,label="Photoélectrique")
plt.plot(Energie,CP_nuc,label="Creation paire nuc")
plt.plot(Energie,CP_el,label="Creation paire électronique")

## verification des donnée avec un dictionnaire
dico = {"Energie": Energie,"Diffusion Compton": Diff_c,"Effet Photoelectrique": Photoel,"Creation de paire nucleaire": CP_nuc,"Creation de paire electronique": CP_el}

for variable, valeurs in dico.items():
    print("Valeurs de", variable, ":", ", ".join(map(str,valeurs))) #on utilise map pour changer les valeurs en string puis ", ".join pour les concatener avec le caractere ", " entre chaque valeur
    print()

plt.legend()
plt.show()

    

book.save('bdd_photons.xlsx')