# -*- coding: utf-8 -*-
"""
Created on Thu Feb 22 20:49:20 2024

@author: mathi
"""
###############
#Import des bibliotheque
###############
from tkinter import *
from tkinter import ttk #pour combobox
import matplotlib.pyplot as plt
import numpy as np

# Importation des outils matplotlib pour créer zone graphique dans fenêtre tkinter 
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import openpyxl as op

##############################################################################
# Fonctions pour convertir les valeurs du curseur en valeurs réelles
def scale_to_realmin(value):
    return value * 1e-4

def scale_to_realmax(value):
    return value * 1e2
##############################################################################
# Fonction pour mettre à jour l'affichage de la valeur réelle de tau
def update_value_tau(value, is_max=False):
    if is_max:
        real_value = scale_to_realmax(curseur_tau_max.get())
        label_valeur_tau_max.config(text=f"     Valeur maximal de tau : {real_value:.4e}     ")
    else:
        real_value = scale_to_realmin(curseur_tau_min.get())
        label_valeur_tau_min.config(text=f"     Valeur minimal de tau : {real_value:.4e}     ")
        
# Fonction pour mettre à jour l'affichage de la valeur réelle de l'energie
def update_value_nrj(value, is_max=False):
    if is_max:
        real_value = scale_to_realmax(curseur_nrj_max.get())
        label_valeur_nrj_max.config(text=f"  Valeur maximal de l'energie : {real_value:.4e}  ")
    else:
        real_value = scale_to_realmin(curseur_nrj_min.get())
        label_valeur_nrj_min.config(text=f"  Valeur minimal de l'energie : {real_value:.4e}  ")
##############################################################################
#fonction qui recupere le choix dans la combobox        
def choix_mat_fct(event):
    global mat_choisi
    mat_choisi = Menu_deroulant.get()
    print("Élément sélectionné :", mat_choisi)
##############################################################################    
def trace():
    # Lecture des datas
    book = op.load_workbook('bdd_photons_all_datas.xlsx')
    sheet = book.get_sheet_by_name(mat_choisi)

    Energie = []
    Diff_ela = []
    Diff_c = []
    Photoel = []
    CP_nuc = []
    CP_el = []
    Tot_w_ela = []
    Tot_wo_ela = []

    for i in range(4,sheet.max_row+1):                                  #on sait que les données commence a la 4ieme ligne
        Energie.append(float(sheet.cell(row = i, column = 1).value))
        Diff_ela.append(float(sheet.cell(row = i, column = 2).value))
        Diff_c.append(float(sheet.cell(row = i, column = 3).value))
        Photoel.append(float(sheet.cell(row = i, column = 4).value))
        CP_nuc.append(float(sheet.cell(row = i, column = 5).value))
        CP_el.append(float(sheet.cell(row = i, column = 6).value))
        Tot_w_ela.append(float(sheet.cell(row = i, column = 7).value))
        Tot_wo_ela.append(float(sheet.cell(row = i, column = 8).value))

    fig.clear() #On efface la zone graphique
    fig.canvas.draw() #On efface le canvas vide
    ax = fig.add_subplot(1,1,1) #graphe pleine page du canvas    
    ax.legend()
    ax.set_title(f"Evolution coeff atténuation massique pour {mat_choisi}")
    ax.set_xlabel("Energie")
    ax.set_ylabel("tau(cm2/g)")
    ax.set_xscale('log')
    ax.set_yscale('log')
    
    ax.set_xlim(scale_to_realmin(curseur_nrj_min.get()),scale_to_realmax(curseur_nrj_max.get()))
    ax.set_ylim(scale_to_realmin(curseur_tau_min.get()),scale_to_realmax(curseur_tau_max.get()))
    
    ax.grid(which = "major", axis='y', linestyle = '--')
    ax.grid(which = "both", axis='x', linestyle = '--')
    
    if PE_ctrl.get() == 1:
        ax.plot(Energie,Photoel,label="Effet photoélectrique")
    
    if Ray_ctrl.get() == 1:
        ax.plot(Energie,Diff_ela,label="Diffusion Rayleigh")
        
    if Comp_ctrl.get() == 1:
        ax.plot(Energie,Diff_c,label="Effet Compton")
        
    if CPn_ctrl.get() == 1:
        ax.plot(Energie,CP_nuc,label="Création de paire nucléaire")
        
    if CPe_ctrl.get() == 1:
        ax.plot(Energie,CP_el,label="Création de paire électronique")
        
    if TotSansRay_ctrl.get() == 1:
        ax.plot(Energie,Tot_wo_ela,label="Atténuation sans diffusion Rayleigh")
            
    if TotAvecRay_ctrl.get() == 1:
        ax.plot(Energie,Tot_w_ela,label="Atténuation avec diffusion Rayleigh")
        
    ax.legend()
    fig.canvas.draw()

    book.save('bdd_photons_all_datas.xlsx')
##############################################################################
def reset():
    fig.clear()
    fig.canvas.draw()

##############
#Programme principal
##############

mat_choisi = "Aluminium"

#création d'une fenêtre tkinter
fenetre = Tk()
#fenetre.geometry("800x500")

#Titre à la fenêtre
fenetre.title("interactions des photon dans la matière")

#creation menu deroulant
Choix_mat = Label(fenetre, text="Choisir un élément Chimique :")
list_materiaux = ["Aluminium", "Plomb", "Cobalt", "Cuivre"]
Menu_deroulant = ttk.Combobox(fenetre, values=list_materiaux)
Menu_deroulant.current(0)

#utilisation menu deroulant
Menu_deroulant.bind("<<ComboboxSelected>>", choix_mat_fct)

#placement menu deroulant
Choix_mat.grid(row=0, column=0, sticky='w',padx=50)
Menu_deroulant.grid(row=0, column=1, sticky='w')

#creation variables de controle
PE_ctrl = IntVar()
Ray_ctrl = IntVar()
Comp_ctrl = IntVar()
CPn_ctrl = IntVar()
CPe_ctrl = IntVar()
TotSansRay_ctrl = IntVar()
TotAvecRay_ctrl = IntVar()

PE_ctrl.set(1)
Comp_ctrl.set(1)
CPn_ctrl.set(1)
CPe_ctrl.set(1)
TotSansRay_ctrl.set(1)

#Creation du cadre selec courbe
cadre_courbe = Frame(fenetre, borderwidth=2, relief="raised")
titre_cadre1 = Label(fenetre,text="Cocher les courbes souhaitées :")
titre_cadre1.grid(row=1,sticky="nw",padx=25,pady=5)

#Creation des cases a cocher
PE_chk = Checkbutton(cadre_courbe, text="Effet photoélectrique", variable=PE_ctrl)
Ray_chk = Checkbutton(cadre_courbe, text="Diffusion Rayleigh", variable=Ray_ctrl)
Comp_chk = Checkbutton(cadre_courbe, text="Effet Compton", variable=Comp_ctrl)
CPn_chk = Checkbutton(cadre_courbe, text="Création de paire nucleaire", variable=CPn_ctrl)
CPe_chk = Checkbutton(cadre_courbe, text="Création de paire electronique", variable=CPe_ctrl)
TotSansRay_chk = Checkbutton(cadre_courbe, text="Atténuation sans diffusion Rayleigh", variable=TotSansRay_ctrl)
TotAvecRay_chk = Checkbutton(cadre_courbe, text="Atténuation avec diffusion Rayleigh", variable=TotAvecRay_ctrl)

#Placement des cases à cocher dans le cadre selec courbe
Label(cadre_courbe).grid(row=1)
PE_chk.grid(row=2, sticky="w")
Ray_chk.grid(row=3, sticky="w")
Comp_chk.grid(row=4, sticky="w")
CPn_chk.grid(row=5, sticky="w")
CPe_chk.grid(row=6, sticky="w")
TotSansRay_chk.grid(row=7, sticky="w")
TotAvecRay_chk.grid(row=8, sticky="w")
Label(cadre_courbe).grid(row=9)

#Placement du cadre selec courbe
cadre_courbe.grid(row=1, column=0, padx=10, pady=10)

#creation cadre val tau
cadre_tau = Frame(fenetre, borderwidth=2, relief="raised")
titre_cadre2 = Label(fenetre,text="Choisi les valeurs extremum de tau :")
titre_cadre2.grid(row=2,sticky="nw",padx=25,pady=5)

#creation des sliders tau
curseur_tau_min = Scale(cadre_tau, from_=1, to=100, resolution=0.01, orient=HORIZONTAL, command=lambda value: update_value_tau(value, is_max=False))
label_valeur_tau_min = Label(cadre_tau, text="Valeur minimal de tau : ")
curseur_tau_min.set(1)
curseur_tau_min.pack()
label_valeur_tau_min.pack()

curseur_tau_max = Scale(cadre_tau, from_=1, to=100, resolution=0.01, orient=HORIZONTAL, command=lambda value: update_value_tau(value, is_max=True))
label_valeur_tau_max = Label(cadre_tau, text="Valeur minimal de tau : ")
curseur_tau_max.set(100)
curseur_tau_max.pack()
label_valeur_tau_max.pack()

update_value_tau(curseur_tau_min.get(), is_max=False)
update_value_tau(curseur_tau_max.get(), is_max=True)

#placement du cadre tau
cadre_tau.grid(row=2, column=0, padx=10, pady=10)

#creation cadre val energie
cadre_nrj = Frame(fenetre, borderwidth=2, relief="raised")
titre_cadre3 = Label(fenetre,text="Choisi les valeurs extremum de l'energie :")
titre_cadre3.grid(row=2,column=1,sticky="nw",padx=15,pady=5)

#creation des sliders
curseur_nrj_min = Scale(cadre_nrj, from_=1, to=100, resolution=0.01, orient=HORIZONTAL, command=lambda value: update_value_nrj(value, is_max=False))
label_valeur_nrj_min = Label(cadre_nrj, text="Valeur minimal de l'énergie : ")
curseur_nrj_min.set(10)
curseur_nrj_min.pack()
label_valeur_nrj_min.pack()

curseur_nrj_max = Scale(cadre_nrj, from_=1, to=100, resolution=0.01, orient=HORIZONTAL, command=lambda value: update_value_nrj(value, is_max=True))
label_valeur_nrj_max = Label(cadre_nrj, text="Valeur minimal de l'énergie : ")
curseur_nrj_max.set(100)
curseur_nrj_max.pack()
label_valeur_nrj_max.pack()

update_value_nrj(curseur_nrj_min.get(), is_max=False)
update_value_nrj(curseur_nrj_max.get(), is_max=True)

#placement du cadre energie
cadre_nrj.grid(row=2, column=1, padx=10, pady=10)

#Canva
fig = Figure(figsize=(6, 6))
graph = FigureCanvasTkAgg(fig, master=fenetre) 

canvas = graph.get_tk_widget() #Ajoute une zone graphe

#bouton fonction
Tracer_courbe_B = Button(fenetre, text="Tracer les courbes",command=trace)

Reinitialis_B = Button(fenetre, text="Réinitialiser",command=reset)

Quitter_B = Button(fenetre, text="Quitter",command=fenetre.destroy)

#placement canva + dernier boutons:
Tracer_courbe_B.grid(row=3, column=0)
Reinitialis_B.grid(row=3,column=1)
Quitter_B.grid(row=3,column=2)

canvas.grid(row=1, column=3, rowspan =3, padx =10, pady =10)

#Détection action souris/clavier
fenetre.mainloop()




