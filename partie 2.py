# -*- coding: utf-8 -*-
"""
Created on Thu Feb 22 17:54:53 2024

@author: mathi
"""
###############
#Import des bibliotheque
###############
from tkinter import *
import matplotlib.pyplot as plt
import numpy as np

# Importation des outils matplotlib pour créer zone graphique dans fenêtre tkinter 
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import openpyxl as op

############################################################################## 
def trace_alu():
    # Lecture des datas
    book = op.load_workbook('bdd_photons.xlsx')
    sheet = book.get_sheet_by_name("Aluminium")

    Energie = []
    Diff_c = []
    Photoel = []
    CP_nuc = []
    CP_el = []

    for i in range(4,sheet.max_row+1):                                  #on sait que les données commence a la 4ieme ligne
        Energie.append(float(sheet.cell(row = i, column = 1).value))
        Diff_c.append(float(sheet.cell(row = i, column = 2).value))
        Photoel.append(float(sheet.cell(row = i, column = 3).value))
        CP_nuc.append(float(sheet.cell(row = i, column = 4).value))
        CP_el.append(float(sheet.cell(row = i, column = 5).value))

    
    fig.clear() #On efface la zone graphique
    fig.canvas.draw() #On efface le canvas vide
    ax = fig.add_subplot(1,1,1) #graphe pleine page du canvas    
    ax.legend()
    ax.set_title("Evolution coeff atténuation massique de l'Aluminium")
    ax.set_xlabel("Energie")
    ax.set_ylabel("tau(cm2/g)")
    ax.set_xscale('log')
    ax.set_yscale('log')
    
    ax.set_xlim(min(Energie),max(Energie))
    ax.set_ylim(min(Photoel),max(Photoel))
    
    ax.grid(which = "major", axis='y', linestyle = '--')
    ax.grid(which = "both", axis='x', linestyle = '--')
    
    if PE_ctrl.get() == 1:                  #on recupere les variables de controle et on fait les plots en conséquences
        ax.plot(Energie,Photoel,label="Effet photoélectrique")
        
    if Comp_ctrl.get() == 1:
        ax.plot(Energie,Diff_c,label="Effet Compton")
        
    if CP_ctrl.get() == 1:
        ax.plot(Energie,CP_nuc,label="Création de paire nucléaire")
        ax.plot(Energie,CP_el,label="Création de paire électronique")
    ax.legend()
    fig.canvas.draw()

    book.save('bdd_photons.xlsx')

##############################################################################
def effacer():
    fig.clear()
    fig.canvas.draw()
    
##############
#Programme principal
##############

#création d'une fenêtre tkinter
fenetre = Tk()

#Titre à la fenêtre
fenetre.title("interactions photon-matière")

#creation variables de controle
PE_ctrl = IntVar()
Comp_ctrl = IntVar()
CP_ctrl = IntVar()

#Création des cases a cocher
PE_chk = Checkbutton(fenetre, text="Effet photoélectrique", variable=PE_ctrl, command=trace_alu)
Comp_chk = Checkbutton(fenetre, text="Effet Compton", variable=Comp_ctrl, command=trace_alu)
CP_chk = Checkbutton(fenetre, text="Création de paire", variable=CP_ctrl, command=trace_alu)


fig = Figure(figsize=(6, 6))
graph = FigureCanvasTkAgg(fig, master=fenetre) 

canvas = graph.get_tk_widget() #Ajoute une zone graphe

Efface_B = Button(fenetre, text="Effacer",command=effacer)

Quitter_B = Button(fenetre, text="Quitter",command=fenetre.destroy)

#Mise en page avec grid:
PE_chk.grid(row=1, sticky='w', padx=10, pady=5)  # sticky='w' aligne à gauche
Comp_chk.grid(row=2, sticky='w', padx=10, pady=5)  # sticky='w' aligne à gauche
CP_chk.grid(row=3, sticky='w', padx=10, pady=5)  # sticky='w' aligne à gauche

Efface_B.grid(row=4,column=0)
Quitter_B.grid(row=4,column=1)

canvas.grid(row=1, column=2, rowspan =4, padx =10, pady =10)


#Détection action souris/clavier
fenetre.mainloop()