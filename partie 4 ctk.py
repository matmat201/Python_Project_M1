# -*- coding: utf-8 -*-
"""
Created on Fri Feb 23 16:50:32 2024

@author: mathi
"""
###############
#Import des bibliotheque
###############
from customtkinter import CTk, CTkLabel, CTkCheckBox, CTkSlider, CTkButton, CTkComboBox, IntVar, CTkFrame, CTkToplevel, StringVar, CTkEntry, CTkRadioButton, CTkTabview, set_appearance_mode, set_widget_scaling, CTkTextbox, CTkSegmentedButton, CTkProgressBar
from CTkToolTip import CTkToolTip

# Importation des outils matplotlib pour créer zone graphique dans fenêtre tkinter 
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import openpyxl as op #lecture de fichier excel

import webbrowser #ouverture de lien internet dans le navigateur

##############################################################################
# Fonctions pour convertir les valeurs du curseur en valeurs réelles
def scale_to_realmin(value):
    return value * 1e-4

def scale_to_realmax(value):
    return value * 1e2
##############################################################################
# Fonction pour mettre à jour l'affichage de la valeur réelle de tau
def update_value_tau(value, is_max=False):
    if is_max:
        real_value = scale_to_realmax(curseur_tau_max.get())
        label_valeur_tau_max.configure(text=f"     Valeur maximal de tau : {real_value:.4e}     ")
    else:
        real_value = scale_to_realmin(curseur_tau_min.get())
        label_valeur_tau_min.configure(text=f"     Valeur minimal de tau : {real_value:.4e}     ")
        
# Fonction pour mettre à jour l'affichage de la valeur réelle de l'energie
def update_value_nrj(value, is_max=False):
    if is_max:
        real_value = scale_to_realmax(curseur_nrj_max.get())
        label_valeur_nrj_max.configure(text=f"  Valeur maximal de l'energie : {real_value:.4e}  ")
    else:
        real_value = scale_to_realmin(curseur_nrj_min.get())
        label_valeur_nrj_min.configure(text=f"  Valeur minimal de l'energie : {real_value:.4e}  ")
##############################################################################
#fonction qui recupere le choix dans la combobox        
def choix_mat_fct(event):
    global mat_choisi
    mat_choisi = Menu_deroulant.get()
    print("Élément sélectionné :", mat_choisi)
##############################################################################    
def recup_data():
    # Lecture des datas
    book = op.load_workbook('bdd_photons_all_datas.xlsx')
    sheet = book[mat_choisi]

    global Energie
    global Diff_ela
    global Diff_c
    global Photoel
    global CP_nuc
    global CP_el
    global Tot_w_ela
    global Tot_wo_ela

    Energie = []
    Diff_ela = []
    Diff_c = []
    Photoel = []
    CP_nuc = []
    CP_el = []
    Tot_w_ela = []
    Tot_wo_ela = []
    
    for i in range(4,sheet.max_row+1):                                  #on sait que les données commence a la 4ieme ligne
        Energie.append(float(sheet.cell(row = i, column = 1).value))
        Diff_ela.append(float(sheet.cell(row = i, column = 2).value))
        Diff_c.append(float(sheet.cell(row = i, column = 3).value))
        Photoel.append(float(sheet.cell(row = i, column = 4).value))
        CP_nuc.append(float(sheet.cell(row = i, column = 5).value))
        CP_el.append(float(sheet.cell(row = i, column = 6).value))
        Tot_w_ela.append(float(sheet.cell(row = i, column = 7).value))
        Tot_wo_ela.append(float(sheet.cell(row = i, column = 8).value))
    
    book.save('bdd_photons_all_datas.xlsx')

def trace():
    if ctrl_conv == 0:
        recup_data()

    fig.clear() #On efface la zone graphique
    fig.canvas.draw() #On efface le canvas vide
    ax = fig.add_subplot(1,1,1) #graphe pleine page du canvas    
    ax.set_title(f"Evolution coeff atténuation massique pour {mat_choisi}")
    ax.set_xlabel("Energie")
    ax.set_ylabel("tau(cm2/g)")
    ax.set_xscale('log')
    ax.set_yscale('log')
    
    if control_modif_nrj == 0:    #on teste la valeur de la variable de controle pour ajuster l'echelle en fonction
        ax.set_xlim(scale_to_realmin(curseur_nrj_min.get()),scale_to_realmax(curseur_nrj_max.get()))
    elif control_modif_nrj == 1:
        ax.set_xlim(float(val_manu_min_nrj.get()),float(val_manu_max_nrj.get()))
        
    if control_modif_tau == 0:
        ax.set_ylim(scale_to_realmin(curseur_tau_min.get()),scale_to_realmax(curseur_tau_max.get()))
    elif control_modif_tau == 1:
        ax.set_ylim(float(val_manu_min_tau.get()),float(val_manu_max_tau.get()))
    
    ax.grid(which = "major", axis='y', linestyle = '--')
    ax.grid(which = "both", axis='x', linestyle = '--')
    
    if PE_ctrl.get() == 1:
        ax.plot(Energie,Photoel,label="Effet photoélectrique")
    
    if Ray_ctrl.get() == 1:
        ax.plot(Energie,Diff_ela,label="Diffusion Rayleigh")
        
    if Comp_ctrl.get() == 1:
        ax.plot(Energie,Diff_c,label="Effet Compton")
        
    if CPn_ctrl.get() == 1:
        ax.plot(Energie,CP_nuc,label="Création de paire nucléaire")
        
    if CPe_ctrl.get() == 1:
        ax.plot(Energie,CP_el,label="Création de paire électronique")
        
    if TotSansRay_ctrl.get() == 1:
        ax.plot(Energie,Tot_wo_ela,label="Atténuation sans diffusion Rayleigh")
            
    if TotAvecRay_ctrl.get() == 1:
        ax.plot(Energie,Tot_w_ela,label="Atténuation avec diffusion Rayleigh")
        
    ax.legend()
    fig.canvas.draw()

##############################################################################
def reset():
    fig.clear()
    ax = fig.add_subplot(1,1,1)
    ax.legend()
    ax.set_title(f"Evolution coeff atténuation massique pour {mat_choisi}")
    ax.set_xlabel("Energie")
    ax.set_ylabel("tau(cm2/g)")
    ax.set_xscale('log')
    ax.set_yscale('log')
    ax.set_xlim(scale_to_realmin(curseur_nrj_min.get()),scale_to_realmax(curseur_nrj_max.get()))
    ax.set_ylim(scale_to_realmin(curseur_tau_min.get()),scale_to_realmax(curseur_tau_max.get()))
    ax.grid(which = "major", axis='y', linestyle = '--')
    ax.grid(which = "both", axis='x', linestyle = '--')
    fig.canvas.draw()
    
    Menu_deroulant.set(list_materiaux[0])
    choix_mat_fct(None)
    
    PE_ctrl.set(1)
    Ray_ctrl.set(0)
    Comp_ctrl.set(1)
    CPn_ctrl.set(1)
    CPe_ctrl.set(1)
    TotSansRay_ctrl.set(0)
    TotAvecRay_ctrl.set(0)
    
    curseur_tau_min.set(1)
    update_value_tau(curseur_tau_min.get(), is_max=False)
    curseur_tau_max.set(100)
    update_value_tau(curseur_tau_max.get(), is_max=True)
    curseur_nrj_min.set(10)
    update_value_nrj(curseur_nrj_min.get(), is_max=False)
    curseur_nrj_max.set(100)
    update_value_nrj(curseur_nrj_min.get(), is_max=True)
    
    Check_ctrl.set(-1)
    
    global control_modif_nrj
    global control_modif_tau
    control_modif_nrj = 0
    control_modif_tau = 0
    
    global ctrl_conv
    ctrl_conv = 0

##############################################################################
def Documentation():
    fen_doc=CTkToplevel(fenetre)
    fen_doc.title("Documentation")
    fen_doc.geometry("800x400")
    fen_doc.grab_set() #Focus au premier plan
    
    Doc = CTkTextbox(fen_doc,width=800,height=400,wrap="word")
    Doc.pack()
    #Doc.insert("0.0", "Documentation :\n\n" + "Lorem ipsum dolor sit amet, consetetur sadipscing elitr, sed diam nonumy eirmod tempor invidunt ut labore et dolore magna aliquyam erat, sed diam voluptua.\n\n" * 20)
    documentation_content = (
        "================================\n"
        "                                 A i d e\n"
        "================================\n\n"
        "Ce programme permet de modeliser l'evolution des coefficients d’atténuation massiques"
        " des photons traversant en fonction de leur energie. De plus il permet de choisir"
        " différents milieux traversée ainsi que different type d'effets physiques tout en"
        " ajustant les paramètres des graphique"
        " D'autre part, il permet a l'utilisateur de recuperer des valeurs precises de ces"
        " coefficient en offrant la possibilité de choisir une énérgie personnalisé. Il offre"
        " aussi la possibilité de convertir ces valeur en section efficace (barn) ainsi que"
        " de sauvegarder les données alors obtenues\n\n"
        "Fonctionnalités principales :\n"
        "- Sélection du matériau\n"
        "- Sélaction du type d'interaction à étudier.\n"
        "- Visualisation des courbes associées aux différents types d'interactions.\n"
        "- Réglage des valeurs extrêmes de tau et de l'énergie pour affiner l'affichage.\n"
        "- Possibilité de saisie manuelle des valeurs de tau et de l'énergie.\n"
        "- Récuperation d'une valeur précise de tau selon les besoin de l'utilisateur\n"
        "- Calcul automatique des sections efficaces correspondante en fonction des paramètres sélectionnés.\n"
        "- Sauvegarde des données et des graphiques générés.\n"
        "- Divers fonction d'interface utilisateur a des fins de personalisation\n\n"
        "Utilisation :\n"
        "1. Sélectionnez le matériau et le type d'interaction a l'aide du menu déroulant et des cases a cocher.\n"
        "(par defaut, la configuration permet de visualiser pour des photon dans l'aluminium, les tau de l'éffet photoélectrique,"
        "l'effet compton, la creation de paire nucléaire et éléctronique ainsi que la somme des effets')\n"
        "2. Ajustez les valeurs extrêmum de tau et d'énergie si nécessaire.\n"
        "3. Appuyez sur le bouton 'Tracer' pour afficher les courbes.\n"
        "4. Appuyez sur le bouton 'Extraction Tau' pour recuperer la valeur de tau pour une énérgie voulue.\n"
        "5. Appuyer sur le bouton 'Calcul Section Efficace' pour convertir la valeur obtenue de tau en barn.\n"
        "6. Si vous souhaitez plutot travailler avec Mu plutôt que Tau appuyer sur le bouton 'Changement unitée'.\n"
        "7. Une fois les valeurs souhaitais obtenue et le graphe tracé avec les données souhaitée il est possible de les enregistrer"
        " dans des fichier externe avec le bouton 'Sauvegarde Externe'.\n"
        "8. Autres fonctions optionnelles => possibilité de réinitialiser les choix a la configuration par defaut avec le bouton correspondant\n"
        "                                                             => possibilité de cocher ou decocher tous les effets pour visualiser rapidement les courbes pour different matériaux\n"
        "9. Il est possible de modifier l'interface graphique : theme de couleur, echelle de zoom, langue. Pour coller au choix de l'utilisateur\n\n"
        "Pour plus d'informations sur le fonctionnement du programme, veuillez vous référer aux liens donnés en crédits"
    )
    Doc.insert("0.0", documentation_content)
    Doc.configure(state="disabled") #on empeche l'utilisateur de modifier le texte
    
##############################################################################
def Val_Tau():
    fen_tau=CTkToplevel(fenetre)
    fen_tau.title("tau")
    fen_tau.grab_set() #Focus au premier plan
    
    global energie
    global tau
    global list_interaction

    #placement menu deroulant
    Choix_mat.grid(row=0, column=0, sticky='w',padx=50)
    Menu_deroulant.grid(row=0, column=1, sticky='w')
    
    energie = StringVar()
    tau = StringVar()
        
    CTkLabel(fen_tau, text="type d'interaction :").grid(row=1,sticky='w',padx=5,pady=5)
    list_interaction = ["Diffusion Rayleigh","Effet Compton","Effet photoélectrique","Création de paire nucleaire","Création de paire electronique","Atténuation avec diffusion Rayleigh","Atténuation sans diffusion Rayleigh"]
    Menu_inter = CTkComboBox(fen_tau, values=list_interaction, command=lambda event: select_inter(Menu_inter.get()))
    Menu_inter.grid(row=1,column=1,padx=5,pady=5)
    Menu_inter.set("choisi une interaction")
    
    CTkLabel(fen_tau, text="energie :").grid(row=2,sticky='w',padx=5,pady=5)
    
    CTkEntry(fen_tau,textvariable=energie).grid(row=2,column=1,padx=5,pady=5)
    
    CTkButton(fen_tau,text="extraction",width=20,command=extraction).grid(row=3,sticky='e',pady=10,column=1,padx=5)
    
    CTkLabel(fen_tau, text="la valeur de tau est :").grid(row=4,sticky='w',padx=5,pady=5)
    CTkEntry(fen_tau,textvariable=tau).grid(row=4,column=1,padx=5,pady=5)
    
##############################################################################
def extraction():
    book = op.load_workbook('bdd_photons_all_datas.xlsx')
    sheet = book[mat_choisi]
    
    energie_entree = float(energie.get())
    
    # Initialisation de la valeur de l'énergie la plus proche et de la différence minimale
    energie_proche = None
    difference_min = float('inf') #on initialise a l'infini comme ça toutes les
    
    for i in range(4, sheet.max_row + 1):
        valeur_energie = float(sheet.cell(row=i, column=1).value)
        difference = abs(energie_entree - valeur_energie)
        
        if difference < difference_min:
            difference_min = difference
            energie_proche = valeur_energie
            
            tau.set("{:.3e}".format(float(sheet.cell(row=i, column=indice).value)))

    book.save('bdd_photons_all_datas.xlsx')
    print(type_inter,energie_entree,energie_proche, tau.get()) #verification
    
##############################################################################
def select_inter(inter):
    global type_inter
    type_inter = inter
    global indice
    indice = list_interaction.index(inter) + 2 #on rajoute 2 car la premiere interaction correspond a la deuxieme colonne dans le excel
    
##############################################################################
def Selection_radioButton():
    if Check_ctrl.get() == 1:
        PE_ctrl.set(1)
        Ray_ctrl.set(1)
        Comp_ctrl.set(1)
        CPn_ctrl.set(1)
        CPe_ctrl.set(1)
        TotSansRay_ctrl.set(1)
        TotAvecRay_ctrl.set(1)
    else:
        PE_ctrl.set(0)
        Ray_ctrl.set(0)
        Comp_ctrl.set(0)
        CPn_ctrl.set(0)
        CPe_ctrl.set(0)
        TotSansRay_ctrl.set(0)
        TotAvecRay_ctrl.set(0)
    
##############################################################################
def change_mode_apparence(event):
    mode = Mode_menu.get()
    set_appearance_mode(mode)
    
def change_zoom(event):
    zoom = Zoom_menu.get()
    zoom = int(zoom.replace("%",""))/100 #set_widget_scaling a besoin d'une valeur entre 0 et 1, on utilise replace pour enlever le %
    set_widget_scaling(zoom)             #puis on divise par 100
    
##############################################################################
def section_eff():
    fen_sec_eff=CTkToplevel(fenetre)
    fen_sec_eff.title("Section efficace")
    fen_sec_eff.grab_set()
    
    if mat_choisi == "Aluminium":
        Text = "de l'"
    else:
        Text = "du "
    
    global sec_eff
    sec_eff = StringVar()
    sec_eff.set("{:.3e}".format(float(float(tau.get())*Mass_atm[mat_choisi]/avogadro/barn)))
    #print(sec_eff.get())
    
    CTkLabel(fen_sec_eff, text=f"La section efficace {Text}{mat_choisi}").grid(row=1,sticky='w',padx=5,pady=5)
    CTkLabel(fen_sec_eff, text=f"pour l'interaction => {list_interaction[indice-2]}").grid(row=2,sticky='w',padx=5,pady=5)
    CTkLabel(fen_sec_eff, text=f"a l'énergie de {energie.get()} MeV").grid(row=3,sticky='w',padx=5,pady=5)
    CTkLabel(fen_sec_eff, text=f"vaut : {sec_eff.get()} barn").grid(row=4,sticky="w",padx=5,pady=5)
    #CTkEntry(fen_sec_eff,textvariable=sec_eff,fg_color="transparent",border_width=0).grid(row=4,padx=(8,0),pady=5)
    
##############################################################################
def fenetre_donnees():
    global fen_sauv
    fen_sauv=CTkToplevel(fenetre)
    fen_sauv.title("Sauvegarde")
    fen_sauv.grab_set()  
    
    global nom_fichier
    nom_fichier = StringVar()
    nom_fichier.set("donnees.txt")
    
    global nom_image
    nom_image = StringVar()
    nom_image.set("image.png")
    
    CTkLabel(fen_sauv, text="Nom du fichier de sauvegarde").grid(row=1,column=1,sticky="w",padx=5, pady=5)
    CTkEntry(fen_sauv, textvariable=nom_fichier).grid(row=1,column=2,sticky="w",padx=5, pady=5)
    CTkButton(fen_sauv, text="Enregistrer",command= sauvegarder_donnees).grid(row=2,columnspan=3,padx=5, pady=5)
    
    CTkLabel(fen_sauv, text="Nom de l'image a enregistrer").grid(row=3,column=1,sticky="w",padx=5, pady=5)
    CTkEntry(fen_sauv, textvariable=nom_image).grid(row=3,column=2,sticky="w",padx=5, pady=5)
    CTkButton(fen_sauv, text="Enregistrer",command= sauvegarder_image).grid(row=4,columnspan=3,padx=5, pady=5)
    
    
def sauvegarder_donnees():
    fichier = open(nom_fichier.get(), "w")    # Ouvrir le fichier en mode écriture
        
    fichier.write("\n\n================================\n")
    fichier.write("***** DONNÉES ENREGISTRÉES *****\n")
    fichier.write("================================\n\n")
        
    fichier.write(f"Matériau : {mat_choisi}\n")
    fichier.write(f"Interaction : {list_interaction[indice-2]}\n")
    fichier.write(f"Énergie : {energie.get()} MeV\n")
    fichier.write(f"Tau : {tau.get()} cm²/g\n")
    fichier.write(f"Section efficace : {sec_eff.get()} Barn\n")
    
    fichier.close()
        
    print("Données enregistrées avec succès dans", nom_fichier.get())
    
def sauvegarder_image():
    fig.savefig(nom_image.get())
    print("Image enregistrée avec succès dans", nom_image.get())
##############################################################################
def saisie_manu_tau():
    global fen_saisi_tau
    fen_saisi_tau=CTkToplevel(fenetre)
    fen_saisi_tau.title("Saisie manuelle tau")
    fen_saisi_tau.grab_set()
    
    global val_manu_max_tau
    global val_manu_min_tau
 
    val_manu_max_tau = StringVar()
    val_manu_max_tau.set("{:.3e}".format(1e+4))
    val_manu_min_tau = StringVar()
    val_manu_min_tau.set("{:.3e}".format(1e-4))
    
    CTkLabel(fen_saisi_tau, text="Atténuation max (cm²/g)").grid(row=2,column=1,sticky="w",padx=5, pady=5)
    CTkEntry(fen_saisi_tau, textvariable=val_manu_max_tau).grid(row=2,column=2,sticky="w",padx=5, pady=5)
    
    CTkLabel(fen_saisi_tau, text="Atténuation min (cm²/g)").grid(row=1,column=1,sticky="w",padx=5, pady=5)
    CTkEntry(fen_saisi_tau, textvariable=val_manu_min_tau).grid(row=1,column=2,sticky="w",padx=5, pady=5)
    
    CTkButton(fen_saisi_tau, text="Valider",command= validation_tau).grid(row=3,columnspan=3,padx=5, pady=5)
    
def saisie_manu_nrj():
    global fen_saisi_nrj
    fen_saisi_nrj=CTkToplevel(fenetre)
    fen_saisi_nrj.title("Saisie manuelle nrj")
    fen_saisi_nrj.grab_set()
    
    global val_manu_max_nrj
    global val_manu_min_nrj

    val_manu_max_nrj = StringVar()
    val_manu_max_nrj.set("{:.3e}".format(1e+4))
    val_manu_min_nrj = StringVar()
    val_manu_min_nrj.set("{:.3e}".format(1e-3))
    
    CTkLabel(fen_saisi_nrj, text="Énergie max (MeV)").grid(row=2,column=1,sticky="w",padx=5, pady=5)
    CTkEntry(fen_saisi_nrj, textvariable=val_manu_max_nrj).grid(row=2,column=2,sticky="w",padx=5, pady=5)
    
    CTkLabel(fen_saisi_nrj, text="Énergie min (MeV)").grid(row=1,column=1,sticky="w",padx=5, pady=5)
    CTkEntry(fen_saisi_nrj, textvariable=val_manu_min_nrj).grid(row=1,column=2,sticky="w",padx=5, pady=5)
    
    CTkButton(fen_saisi_nrj, text="Valider",command= validation_nrj).grid(row=3,columnspan=3,padx=5, pady=5)
    
def validation_tau():
    global control_modif_tau
    control_modif_tau = 1
    fen_saisi_tau.destroy()
    trace()

def validation_nrj():
    global control_modif_nrj
    control_modif_nrj = 1
    fen_saisi_nrj.destroy()
    trace()

##############################################################################
def Text_survol():
    print(tooltip_Unite.get())
    
##############################################################################
def Conversion():
    
    global Diff_ela
    global Diff_c
    global Photoel
    global CP_nuc
    global CP_el
    global Tot_w_ela
    global Tot_wo_ela
    global ctrl_conv
    
    if ctrl_conv == 0:
        Diff_ela = [val_tau * rho[mat_choisi] for val_tau in Diff_ela] #on modifie la liste par comprehension
        Diff_c = [val_tau * rho[mat_choisi] for val_tau in Diff_c]     #avec une variable anonyme qu'on itère sur la liste
        Photoel = [val_tau * rho[mat_choisi] for val_tau in Photoel]
        CP_nuc = [val_tau * rho[mat_choisi] for val_tau in CP_nuc]
        CP_el = [val_tau * rho[mat_choisi] for val_tau in CP_el]
        Tot_w_ela = [val_tau * rho[mat_choisi] for val_tau in Tot_w_ela]
        Tot_wo_ela = [val_tau * rho[mat_choisi] for val_tau in Tot_wo_ela]

        ctrl_conv = 1
        
    elif ctrl_conv == 1:
        Diff_ela = [val_tau / rho[mat_choisi] for val_tau in Diff_ela]
        Diff_c = [val_tau / rho[mat_choisi] for val_tau in Diff_c]
        Photoel = [val_tau / rho[mat_choisi] for val_tau in Photoel]
        CP_nuc = [val_tau / rho[mat_choisi] for val_tau in CP_nuc]
        CP_el = [val_tau / rho[mat_choisi] for val_tau in CP_el]
        Tot_w_ela = [val_tau / rho[mat_choisi] for val_tau in Tot_w_ela]
        Tot_wo_ela = [val_tau / rho[mat_choisi] for val_tau in Tot_wo_ela]

        ctrl_conv = 0   
    
    trace()
    
def credit():
    fen_cred=CTkToplevel(fenetre)
    fen_cred.title("Sauvegarde")
    fen_cred.grab_set()
    
    global link_selec
    link_selec = StringVar()
    
    CTkLabel(fen_cred, text="Dev by Mathis BRONNER-BERENGER\n""during a programming project in\n""first year of nuclear engineering\n""master degree").grid(row=1,padx=5, pady=5)
    Link_B = CTkButton(fen_cred, text="Usefull links\n""↓",text_color_disabled="white")
    Link_B.grid(row=2,padx=5, pady=5)
    Link_B.configure(state="disabled")
    CTkSegmentedButton(fen_cred, values=["Project link","CustomTkinter","CTkToolTip"],command=ouvre_lien,variable=link_selec).grid(row=3,padx=5, pady=5)
    CTkLabel(fen_cred, text="Thank you for using my little app").grid(row=4,padx=5, pady=5)
    progressbar_1 = CTkProgressBar(fen_cred)
    progressbar_1.grid(row=5, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")
    progressbar_1.configure(mode="indeterminnate") #fait en sorte que la barre de progression alterne de gauche a droite
    progressbar_1.start() #lance la progression de la barre
    

def ouvre_lien(event):
    if link_selec.get() == "Project link":
        webbrowser.open_new(git_1)
    if link_selec.get() == "CustomTkinter":
        webbrowser.open_new(git_2)
    if link_selec.get() == "CTkToolTip":
        webbrowser.open_new(git_3)
##############
#Programme principal
##############
control_modif_tau = 0

control_modif_nrj = 0

ctrl_conv = 0

global avogadro
avogadro = 6.022e23

global barn
barn = 1e-24

Mass_atm = {"Aluminium": 26.98, "Plomb": 207.2, "Cobalt": 58.93, "Cuivre": 63.55}

rho = {"Aluminium": 2.698, "Plomb": 11.342, "Cobalt": 8.900, "Cuivre": 8.960}

mat_choisi = "Aluminium"

git_1 = "https://github.com/matmat201/Python_Project_M1"
git_2 = "https://customtkinter.tomschimansky.com"
git_3 = "https://github.com/Akascape/CTkToolTip"

#création d'une fenêtre tkinter
fenetre = CTk()
fenetre.geometry("1050x570")

#Titre à la fenêtre
fenetre.title("interactions des photon dans la matière")

#creation menu deroulant
Choix_mat = CTkLabel(fenetre, text="Choisir un élément Chimique :")
list_materiaux = ["Aluminium", "Plomb", "Cobalt", "Cuivre"]
Menu_deroulant = CTkComboBox(fenetre, values=list_materiaux,command=choix_mat_fct)
Menu_deroulant.set(list_materiaux[0])

#placement menu deroulant
Choix_mat.grid(row=0, column=0, sticky='w',padx=50)
Menu_deroulant.grid(row=0, column=1, sticky='w')

#creation variables de controle
PE_ctrl = IntVar()
Ray_ctrl = IntVar()
Comp_ctrl = IntVar()
CPn_ctrl = IntVar()
CPe_ctrl = IntVar()
TotSansRay_ctrl = IntVar()
TotAvecRay_ctrl = IntVar()

PE_ctrl.set(1)
Comp_ctrl.set(1)
CPn_ctrl.set(1)
CPe_ctrl.set(1)
TotSansRay_ctrl.set(1)

# Définir le poids des lignes et colonnes pour l'expansion automatique
fenetre.grid_rowconfigure(0, weight=1)
fenetre.grid_columnconfigure(0, weight=1)
fenetre.grid_rowconfigure(1, weight=1)
fenetre.grid_columnconfigure(1, weight=1)
fenetre.grid_rowconfigure(2, weight=1)
fenetre.grid_columnconfigure(2, weight=1)
fenetre.grid_rowconfigure(3, weight=1)
fenetre.grid_columnconfigure(3, weight=1)

#Creation du cadre selec courbe
cadre_courbe = CTkFrame(fenetre)
titre_cadre1 = CTkLabel(cadre_courbe, text="Cocher les courbes souhaitées :")
titre_cadre1.grid(row=0,sticky="nw",padx=25,pady=5)

#Creation des cases a cocher
PE_chk = CTkCheckBox(cadre_courbe, text="Effet photoélectrique", variable=PE_ctrl)
Ray_chk = CTkCheckBox(cadre_courbe, text="Diffusion Rayleigh", variable=Ray_ctrl)
Comp_chk = CTkCheckBox(cadre_courbe, text="Effet Compton", variable=Comp_ctrl)
CPn_chk = CTkCheckBox(cadre_courbe, text="Création de paire nucleaire", variable=CPn_ctrl)
CPe_chk = CTkCheckBox(cadre_courbe, text="Création de paire electronique", variable=CPe_ctrl)
TotSansRay_chk = CTkCheckBox(cadre_courbe, text="Atténuation sans diffusion Rayleigh", variable=TotSansRay_ctrl)
TotAvecRay_chk = CTkCheckBox(cadre_courbe, text="Atténuation avec diffusion Rayleigh", variable=TotAvecRay_ctrl)

#Placement des cases à cocher dans le cadre selec courbe
PE_chk.grid(row=1, sticky="w",pady=5)
Ray_chk.grid(row=2, sticky="w",pady=5)
Comp_chk.grid(row=3, sticky="w",pady=5)
CPn_chk.grid(row=4, sticky="w",pady=5)
CPe_chk.grid(row=5, sticky="w",pady=5)
TotSansRay_chk.grid(row=6, sticky="w",pady=5)
TotAvecRay_chk.grid(row=7, sticky="w",pady=5)

#Placement du cadre selec courbe
cadre_courbe.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

#creation cadre val tau
cadre_tau = CTkFrame(fenetre)
titre_cadre2 = CTkLabel(cadre_tau, text="Choisi les valeurs extremum de tau :")
titre_cadre2.pack()

#creation des sliders tau
curseur_tau_min = CTkSlider(cadre_tau, from_=1, to=100, command=lambda value: update_value_tau(value, is_max=False))
label_valeur_tau_min = CTkLabel(cadre_tau, text="Valeur minimal de tau : ")
curseur_tau_min.set(1)
curseur_tau_min.pack()
label_valeur_tau_min.pack()

curseur_tau_max = CTkSlider(cadre_tau, from_=1, to=100, command=lambda value: update_value_tau(value, is_max=True))
label_valeur_tau_max = CTkLabel(cadre_tau, text="Valeur maximal de tau : ")
curseur_tau_max.set(100)
curseur_tau_max.pack()
label_valeur_tau_max.pack()

update_value_tau(curseur_tau_min.get(), is_max=False)
update_value_tau(curseur_tau_max.get(), is_max=True)

bouton_saisie_manu1 = CTkButton(cadre_tau,text="Saisie manuelle",command=saisie_manu_tau)
bouton_saisie_manu1.pack()

#placement du cadre tau
cadre_tau.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

#creation cadre val energie
cadre_nrj = CTkFrame(fenetre)
titre_cadre3 = CTkLabel(cadre_nrj, text="Choisi les valeurs extremum de l'energie :")
titre_cadre3.pack()

#creation des sliders
curseur_nrj_min = CTkSlider(cadre_nrj, from_=1, to=100, command=lambda value: update_value_nrj(value, is_max=False))
label_valeur_nrj_min = CTkLabel(cadre_nrj, text="Valeur minimal de l'énergie : ")
curseur_nrj_min.set(10)
curseur_nrj_min.pack()
label_valeur_nrj_min.pack()

curseur_nrj_max = CTkSlider(cadre_nrj, from_=1, to=100, command=lambda value: update_value_nrj(value, is_max=True))
label_valeur_nrj_max = CTkLabel(cadre_nrj, text="Valeur maximal de l'énergie : ")
curseur_nrj_max.set(100)
curseur_nrj_max.pack()
label_valeur_nrj_max.pack()

update_value_nrj(curseur_nrj_min.get(), is_max=False)
update_value_nrj(curseur_nrj_max.get(), is_max=True)

bouton_saisie_manu2 = CTkButton(cadre_nrj,text="Saisie manuelle",command=saisie_manu_nrj)
bouton_saisie_manu2.pack()

#placement du cadre energie
cadre_nrj.grid(row=2, column=1, padx=10, pady=10, sticky="nsew")

#Canva
fig = Figure(figsize=(6, 6))
graph = FigureCanvasTkAgg(fig, master=fenetre) 

canvas = graph.get_tk_widget() #Ajoute une zone graphe

#bouton fonction
Tracer_courbe_B = CTkButton(fenetre, text="Tracer les courbes",command=trace)
Reinitialis_B = CTkButton(fenetre, text="Réinitialiser",command=reset)
Quitter_B = CTkButton(fenetre, text="Quitter",command=fenetre.destroy)

#placement canva + dernier boutons:
Tracer_courbe_B.grid(row=3, column=0)
Reinitialis_B.grid(row=3,column=1)
Quitter_B.grid(row=3,column=2)

canvas.grid(row=1, column=2, rowspan=2, padx=10, pady=10, sticky="nsew")

#tabview des nouvelles fonctions et des options
Tabview_fonc = CTkTabview(fenetre)
Tabview_fonc.add("Fonction Part 4")
Tabview_fonc.add("Options")
Tabview_fonc.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")


Aide_B = CTkButton(Tabview_fonc.tab("Fonction Part 4"), text="Aide",command=Documentation).grid(row=1,pady=5,padx=(40,0))
Coef_B = CTkButton(Tabview_fonc.tab("Fonction Part 4"), text="Extraction Tau",command=Val_Tau).grid(row=2,pady=5,padx=(40,0))
Section_B = CTkButton(Tabview_fonc.tab("Fonction Part 4"), text="Calcul Section Efficace",command=section_eff).grid(row=3,pady=5,padx=(40,0))
Unite_B = CTkButton(Tabview_fonc.tab("Fonction Part 4"), text="Changement unité",command=Conversion)
Unite_B.grid(row=4,pady=5,padx=(40,0))
Save_B = CTkButton(Tabview_fonc.tab("Fonction Part 4"), text="Sauvegarde Externe",command=fenetre_donnees).grid(row=5,pady=5,padx=(40,0))

tooltip_Unite = CTkToolTip(Unite_B, delay=0.1, message="Ce bouton permet de convertir\n"
                 'la valeur d\'aténuation tau (cm²/g)\n'
                 'en coefficient d\'atténuation linéique mu (cm-1)')

Check_ctrl = IntVar(value=-1)
Coche = CTkRadioButton(Tabview_fonc.tab("Fonction Part 4"), text="Check All",command=Selection_radioButton, variable= Check_ctrl, value=1).grid(row=6,padx=(0,80))
Decoche = CTkRadioButton(Tabview_fonc.tab("Fonction Part 4"), text="Uncheck All",command=Selection_radioButton, variable= Check_ctrl, value=0).grid(row=6,padx=(150,0))


Mode_ent = CTkLabel(Tabview_fonc.tab("Options"),text="Mode d'apparence").grid(row=1,column=1,pady=10,padx=(10,0))
Mode_menu = CTkComboBox(Tabview_fonc.tab("Options"),values=["Light", "Dark", "System"],command=change_mode_apparence)
Mode_menu.grid(row=1,column=2,pady=10,padx=(10,0))
Mode_menu.set("Dark")

Zoom_ent = CTkLabel(Tabview_fonc.tab("Options"),text="Zoom").grid(row=2,column=1,pady=10,padx=(10,0))
Zoom_menu = CTkComboBox(Tabview_fonc.tab("Options"),values=["80%", "90%", "100%", "110%", "120%"],command=change_zoom)
Zoom_menu.grid(row=2,column=2,pady=10,padx=(10,0))
Zoom_menu.set("100")

Language_ent = CTkLabel(Tabview_fonc.tab("Options"),text="Langue").grid(row=3,column=1,pady=10,padx=(10,0))
Language_menu = CTkComboBox(Tabview_fonc.tab("Options"),values=["Français", "WIP", "..."])
Language_menu.grid(row=3,column=2,pady=10,padx=(10,0))
Language_menu.set("Français")

Credit_B = CTkButton(Tabview_fonc.tab("Options"),text="Crédits",command=credit).grid(row=4,columnspan=3,pady=(50,0),padx=(15,0))

trace()

#Détection action souris/clavier
fenetre.mainloop()
